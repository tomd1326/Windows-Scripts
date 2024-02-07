import shutil
import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows

# Function to print messages with timestamps
def print_with_timestamp(message):
    now = datetime.now()
    timestamp = now.strftime("[%Y-%m-%d %H:%M:%S]")
    print(f"{timestamp} {message}")

# Paths
template_folder = r'C:\Users\Tom\OneDrive - Motor Depot\Pricing'
input_folder = r'C:\Users\Tom\OneDrive - Motor Depot\Pricing\Input Files'

# Find the template file
template_files = [f for f in os.listdir(template_folder) if f.startswith('__TEMPLATE Stock and checks') and f.endswith('.xlsx')]
if not template_files:
    print_with_timestamp("No template file found.")
    exit()

template_file = os.path.join(template_folder, template_files[0])
now = datetime.now()
date_time_str = now.strftime("%Y_%m_%d_%H%M")
new_file_name = f'Stock and checks {date_time_str}.xlsx'
destination_file = os.path.join(template_folder, new_file_name)

# Copy and rename the template file
try:
    shutil.copy(template_file, destination_file)
    os.chmod(destination_file, 0o777)
    print_with_timestamp(f"Created a copy of '{template_file}' as '{new_file_name}'.")
except Exception as e:
    print_with_timestamp(f"An error occurred while copying the template file: {str(e)}")
    exit()

if not os.access(destination_file, os.W_OK):
    print_with_timestamp(f"File '{destination_file}' is not writable.")
    exit()

# Criteria for file selection
criteria = {
    'Apex_Stock': {'prefix': 'vehicles-autoedit-', 'extension': 'csv', 'min_size_kb': 700},
    'VA': {'prefix': 'CAP_Stock_Output_', 'extension': 'csv'},
    'AT_Forecourt': {'prefix': '__MASTER_Forecourt_', 'extension': 'csv'},
    'AT_Initial': {'prefix': '__MASTER_Initial_AT', 'extension': 'csv'},
    'Loc_Hist': {'prefix': 'vehicles-location-history', 'extension': 'csv'},
    '0%_Stock': {'prefix': 'vehicles-autoedit-', 'extension': 'csv', 'max_size_kb': 200},
    'C2P': {'prefix': 'Cars to price', 'extension': 'csv'}
}

# Function to check if file size is valid
def is_size_valid(file_path, min_size_kb=None, max_size_kb=None):
    file_size_kb = os.path.getsize(file_path) / 1024
    if min_size_kb and file_size_kb < min_size_kb:
        return False
    if max_size_kb and file_size_kb > max_size_kb:
        return False
    return True

# Function to clear an Excel sheet
def clear_sheet(worksheet):
    for row in worksheet.iter_rows():
        for cell in row:
            cell.value = None

# Function to copy CSV data to Excel sheet
def copy_data_to_excel(sheet_name, file_path, workbook):
    encoding_to_try = ['utf-8', 'ISO-8859-1', 'cp1252']  # List of encodings to try

    for encoding in encoding_to_try:
        try:
            data = pd.read_csv(file_path, encoding=encoding, dtype=str)
            break  # Break the loop if reading is successful
        except UnicodeDecodeError:
            continue  # Try the next encoding

    else:  # If all encodings fail
        print_with_timestamp(f"Error decoding file '{file_path}' with all tried encodings.")
        return

    # Convert specified numeric columns to numbers
    numeric_columns = ['Retail price', 'Price position', 'Retail valuation', 'Auto Trader Retail Rating']  # Add other numeric columns as needed
    for col in numeric_columns:
        if col in data.columns:
            data[col] = pd.to_numeric(data[col], errors='coerce')

    worksheet = workbook[sheet_name]
    clear_sheet(worksheet)  # Clear the sheet

    # Writing data to the worksheet
    for r_idx, row in enumerate(data.itertuples(index=False, name=None), start=2):  # start=2 to skip header row
        for c_idx, value in enumerate(row, start=1):
            worksheet.cell(row=r_idx, column=c_idx, value=value)

    # Write headers
    for c_idx, header in enumerate(data.columns, start=1):
        worksheet.cell(row=1, column=c_idx, value=header)

    print_with_timestamp(f"Data from '{file_path}' copied to '{sheet_name}'. Rows copied: {len(data)}")


# Main script logic
try:
    workbook = load_workbook(destination_file)

    for sheet, crit in criteria.items():
        found = False
        for file in os.listdir(input_folder):
            if file.startswith(crit['prefix']) and file.endswith('.csv'):
                file_path = os.path.join(input_folder, file)
                if is_size_valid(file_path, crit.get('min_size_kb'), crit.get('max_size_kb')):
                    copy_data_to_excel(sheet, file_path, workbook)
                    found = True
                    break
        if not found:
            print_with_timestamp(f"No suitable CSV file found for sheet '{sheet}'.")

    if 'OUTPUT' in workbook.sheetnames:
        workbook.active = workbook.sheetnames.index('OUTPUT')

    # Save the workbook before closing
    workbook.save(destination_file)
    print_with_timestamp(f"Updated Excel file saved as '{destination_file}'.")

except Exception as e:
    print_with_timestamp(f"An error occurred: {e}")
