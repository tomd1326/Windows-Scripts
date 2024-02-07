import openpyxl
import os
from datetime import datetime
import glob
from openpyxl.workbook.properties import CalcProperties

# Define output_headers outside of the main function
output_headers = [
    "Supp", "VRM", "Make", "Model", "F", "Yr", "Description", "Dr", "Spec", "Spec £", "Yr Pl", "Mileage", "Colour", "A", "IF4C",
    "By", "Status", "Sold", "CAP Cl Mth", "CAP Cl Live", "Cost", "SIV", "Current Price",
    "DIS", "Days Live", "0% Days", "AT Price", "1st AT", "CAP Ret Live",
    "RR", "AT Perf", "AT Last Change", "Chg Date", "Chg Days", "Total change", "Review Due", "Last Check", "Recon", "Net Margin", "7 day views", "Stock", "Loc", "Review Cat", "Price Cat", "Img", "CAPID", "To price", "Sugg"
]

def normalize_header(header):
    if header is None:
        return None
    return ' '.join(header.replace("\n", " ").replace("\r", " ").split())

def find_column_indices(sheet, headers):
    header_to_index = {normalize_header(cell.value): cell.column for cell in sheet[1]}
    return {header: header_to_index.get(header) for header in headers if header in header_to_index}

def replace_formulas_with_values(sheet, values_sheet, column_indices):
    for row in range(2, sheet.max_row + 1):
        for col_header, col_index in column_indices.items():
            if col_index is None:
                continue
            
            original_cell = sheet.cell(row=row, column=col_index)
            value_cell = values_sheet.cell(row=row, column=col_index)
            original_cell.value = value_cell.value

def process_output_sheet(workbook, values_workbook, sheet_name, headers):
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        values_sheet = values_workbook[sheet_name]

        column_indices = find_column_indices(sheet, headers)

        replace_formulas_with_values(sheet, values_sheet, column_indices)

def process_schedule_sheet(workbook, values_workbook, sheet_name):
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        values_sheet = values_workbook[sheet_name]

        # Unmerge all cells
        for merged_cell_range in list(sheet.merged_cells.ranges):
            sheet.unmerge_cells(str(merged_cell_range))

        # Update cell values
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=sheet.max_column):
            for cell in row:
                value_cell = values_sheet.cell(row=cell.row, column=cell.column)
                cell.value = value_cell.value

def custom_sort_key(row):
    make = row[output_headers.index("Make")]
    model = row[output_headers.index("Model")]
    yr = row[output_headers.index("Yr")]
    description = row[output_headers.index("Description")]
    dr = row[output_headers.index("Dr")]
    try:
        cap_cl_live = float(row[output_headers.index("CAP Cl Live")])
    except ValueError:
        cap_cl_live = 0.0  # Set a default value if conversion fails

    return (make, model, yr, description, dr, cap_cl_live)

def main():
    # Get today's date in yyyy_mm_dd format
    today_date = datetime.now().strftime("%Y_%m_%d")

    # Define the directory where the files are located
    directory = r"C:\Users\Tom\OneDrive - Motor Depot\Pricing"

    # Find the file that matches the pattern with today's date
    file_pattern = os.path.join(directory, f"Stock and checks {today_date}_*.xlsx")
    matching_files = glob.glob(file_pattern)

    if not matching_files:
        print("No file found for today's date.")
        return

    # Assuming you want to process the first file found
    file_path = matching_files[0]

    print(f"Starting processing of workbook '{file_path}'...")

    print("Loading workbooks...")

    # Initialize start_time at the beginning of the operation
    start_time = datetime.now()

    # Your existing code for loading workbooks
    workbook = openpyxl.load_workbook(file_path, data_only=False)
    values_workbook = openpyxl.load_workbook(file_path, data_only=True)

    # Trigger recalculation
    workbook.calculation = "automatic"


    # Your existing code for calculating elapsed time
    end_time = datetime.now()
    elapsed_time = end_time - start_time
    print(f"Loading workbooks completed. Elapsed time: {elapsed_time}")

    print("Processing the 'OUTPUT' sheet...")
    start_time = datetime.now()

    output_sheet = workbook["OUTPUT"]
    print(f"Number of rows in 'OUTPUT' sheet: {output_sheet.max_row}")

    process_output_sheet(workbook, values_workbook, "OUTPUT", output_headers)

    end_time = datetime.now()
    elapsed_time = end_time - start_time
    print(f"Processing 'OUTPUT' sheet completed. Elapsed time: {elapsed_time}")

    print("Processing the 'Schedule' sheet...")
    start_time = datetime.now()
    process_schedule_sheet(workbook, values_workbook, "Schedule")
    end_time = datetime.now()
    elapsed_time = end_time - start_time
    print(f"Processing 'Schedule' sheet completed. Elapsed time: {elapsed_time}")

    # Deleting specified worksheets
    sheets_to_delete = ["MASTER_work", "CAP_Models", "Comm", "Lookups", "Pricing_Formula", "Apex_Stock", "AT_Forecourt", "AT_Initial", "Loc_Hist", "0%_Stock", "C2P", "Price_changes", "RWA", "VA"]
    for sheet in sheets_to_delete:
        if sheet in workbook.sheetnames:
            start_time = datetime.now()
            del workbook[sheet]
            end_time = datetime.now()
            elapsed_time = end_time - start_time
            print(f"Deleted sheet: {sheet}. Elapsed time: {elapsed_time}")

    # Save the workbook with a new name
    base_name, extension = os.path.splitext(file_path)
    new_file_path = f"{base_name}_values{extension}"
    workbook.save(new_file_path)  # Save the workbook with the new name
    print(f"Workbook saved as '{new_file_path}'")

    # 1) Delete all rows on the OUTPUT sheet where the first column ("Supp") is 0
    output_sheet = workbook["OUTPUT"]
    rows_to_delete = []
    for row in output_sheet.iter_rows(min_row=2, max_row=output_sheet.max_row, max_col=1):
        if row[0].value == 0:
            rows_to_delete.append(row[0].row)

    for row in reversed(rows_to_delete):
        output_sheet.delete_rows(row)

    # 2) Sort the OUTPUT sheet by: Make, Model, Yr, Description, Dr, CAP Cl Live
    start_time = datetime.now()
    
    output_sheet = workbook["OUTPUT"]
    output_data = []
    for row in output_sheet.iter_rows(min_row=2, values_only=True):
        output_data.append(row)

    try:
        output_data.sort(key=custom_sort_key)
    except TypeError as e:
        print(f"Sorting 'OUTPUT' sheet failed: {e}")

    for i, row in enumerate(output_data, start=2):
        for j, value in enumerate(row, start=1):
            output_sheet.cell(row=i, column=j, value=value)
    
    end_time = datetime.now()
    elapsed_time = end_time - start_time
    print(f"Sorting 'OUTPUT' sheet completed. Elapsed time: {elapsed_time}")

if __name__ == "__main__":
    main()
